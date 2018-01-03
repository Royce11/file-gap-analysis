import openpyxl
from openpyxl.styles import NamedStyle,Font,Alignment

unprocessedWB_out = '/home/osboxes/shared-windows10/UnprocessedDates.xlsx'
missingWB_out = '/home/osboxes/shared-windows10/MissingDates.xlsx'

def addHeaderStyle(sheetTitle):
    missingWB = openpyxl.load_workbook(missingWB_out)
    unprocessedWB = openpyxl.load_workbook(unprocessedWB_out)
    currentSheet1 = missingWB.get_sheet_by_name(sheetTitle)
    currentSheet2 = unprocessedWB.get_sheet_by_name(sheetTitle)
    if 'red_bold' not in missingWB._named_styles.names:
        red_bold = NamedStyle(name="red_bold")
        red_bold.font = Font(color='00FF0000', bold=True)
        missingWB.add_named_style(red_bold)

    if 'red_bold' not in unprocessedWB._named_styles.names:
        red_bold = NamedStyle(name="red_bold")
        red_bold.font = Font(color='00FF0000', bold=True)
        unprocessedWB.add_named_style(red_bold)

    for cell in currentSheet1["1:1"]:
        cell.style = 'red_bold'
    for cell in currentSheet2["1:1"]:
        cell.style = 'red_bold'

    missingWB.save(missingWB_out)
    unprocessedWB.save(unprocessedWB_out)

def initExcelSheet(sheetTitle,fieldnames,vendorCount):
    missingWB = openpyxl.load_workbook(missingWB_out)
    unprocessedWB = openpyxl.load_workbook(unprocessedWB_out)
    missingWB.create_sheet(sheetTitle,vendorCount)
    unprocessedWB.create_sheet(sheetTitle,vendorCount)
    currentSheet1 = missingWB.get_sheet_by_name(sheetTitle)
    currentSheet2 = unprocessedWB.get_sheet_by_name(sheetTitle)

    i=0
    while i < len(fieldnames):
        currentSheet1.cell(row=1,column=i+1,value=fieldnames[i])

        currentSheet2.cell(row=1,column=i+1,value=fieldnames[i])
        i=i+1
    missingWB.save(missingWB_out)
    unprocessedWB.save(unprocessedWB_out)

    addHeaderStyle(sheetTitle)