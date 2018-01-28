import json
import openpyxl
import os
from com.reporting import RentrakHandler,ExcelUtilities,HPEHandler,ComscoreHandler,SharethisHandler
import sys
from datetime import datetime,date

unprocessedWB_out = '/home/osboxes/shared-windows10/UnprocessedDates.xlsx'
missingWB_out = '/home/osboxes/shared-windows10/MissingDates.xlsx'

def initExcelWB():

    # missingWorkbook = openpyxl.load_workbook(missingWB_out)
    # unprocessedWorkbook = openpyxl.load_workbook(unprocessedWB_out)
    #
    # for i in missingWorkbook.worksheets:
    #     missingWorkbook.remove_sheet(i)
    #
    # for i in unprocessedWorkbook.worksheets:
    #     unprocessedWorkbook.remove_sheet(i)

    missingWorkbook = openpyxl.Workbook()
    unprocessedWorkbook = openpyxl.Workbook()

    missingWorkbook.save(missingWB_out)
    unprocessedWorkbook.save(unprocessedWB_out)

initExcelWB()


def getHandler(vendors, **keywords):
    vendor = keywords.get('vendor')

    if len(sys.argv)>=2:
        inputStartDate = datetime.date(datetime.strptime(sys.argv[1],"%Y-%m-%d"))
    else:
        inputStartDate = date(2016,1,1)

    if len(sys.argv)>=3:
        inputEndDate = datetime.date(datetime.strptime(sys.argv[2],"%Y-%m-%d"))
    else:
        inputEndDate = date.today()

    if vendor == 'rentrak':
        RentrakHandler.processExecute(vendors,inputStartDate,inputEndDate,**keywords)
    if vendor == 'hpe':
        HPEHandler.processExecute(vendors,inputStartDate,inputEndDate,**keywords)
    if vendor == 'comscore':
        ComscoreHandler.processExecute(vendors,inputStartDate,inputEndDate,**keywords)
        # ComscoreHandler.ComscoreHandler.processExecute(vendors,inputStartDate,inputEndDate,**keywords)
    if vendor == 'sharethis':
        SharethisHandler.processExecute(vendors,inputStartDate,inputEndDate,**keywords)



# pathname = '/home/osboxes/shared-windows10/Configs'

# for filename in glob.glob(glob.escape(pathname)+'/.*.json'):
#     with open(filename,'r') as json_data:
#         rule_dict = json.load(json_data)

for file in os.listdir("/home/osboxes/shared-windows10/Configs/Playground/"):
    if file.endswith(".json"):
        filename = os.path.join("/home/osboxes/shared-windows10/Configs/Playground/", file)
        jsonObj = open(filename,'r')
        rule_dict = json.load(jsonObj)

        for vendorCount, vendor in enumerate(rule_dict):
            ExcelUtilities.initExcelSheet(vendor.get('vendor'),vendor.get('fieldnames'),vendorCount)
            keywords = {'vendor': vendor.get('vendor'),'fieldnames' : vendor.get('fieldnames')}

            getHandler(vendor,**keywords)