from com.reporting import S3Utilities,GeneralUtils,ExcelUtilities
from datetime import datetime,date
from isoweek import Week
from collections import OrderedDict
from openpyxl.styles import Alignment

unprocessedWB_out = '/home/osboxes/shared-windows10/UnprocessedDates.xlsx'
missingWB_out = '/home/osboxes/shared-windows10/MissingDates.xlsx'
client = S3Utilities.getS3Client()

def hpeRawFileHandler(response,dateRegex):
    generalFilePatternDict = dict()

    for fullFileName in response:
        dateAndSourceDict = GeneralUtils.extractDateAndDatasource(dateRegex,fullFileName)
        if dateAndSourceDict:
            for source,date in dateAndSourceDict.items():
                generalFilePatternDict.setdefault(str(source),[]).append(GeneralUtils.getFormattedDate(date))

    return generalFilePatternDict

def hpeCleanFileHandler(response):
    generalFilePattern=[]
    FilePattern = GeneralUtils.namedtuple_with_defaults('FilePattern','general')
    if "CommonPrefixes" in response.keys():
        for prefixDictElement in response.get('CommonPrefixes'):
            generalFilePattern.append(GeneralUtils.getFormattedDate(prefixDictElement.get('Prefix').split('/')[5].strip("dt=")))
    return FilePattern(set(generalFilePattern))

def hpeRowWriter(currentSheet,datesSet,startRow,**keywords):
    tempDict = OrderedDict()

    currentSheet._current_row = startRow

    if len(datesSet):
        for dateElement in datesSet:
            tempDict.setdefault(str(dateElement.year) + '-' + str(dateElement.month),[]).append(dateElement.day)
        for key,value in tempDict.items():
            keywords['year-month'] = key
            keywords['dates'] = str(value)
            keywords['count'] = len(value)
            #vendor	datasource	cadence	type	country	year-month	dates	count
            fieldNamesDict = {'A': keywords.get('datasource'),
                              'B' : keywords.get('cadence'),
                              'C' : keywords.get('type'),
                              'D' : keywords.get('country'),
                              'E' : keywords.get('year-month'),
                              'F' : keywords.get('dates'),
                              'G' : keywords.get('count')}
            currentSheet.append(fieldNamesDict)

        endRow = currentSheet.max_row
        if currentSheet.max_row > 1:
            for colNum in range(1,5):
                currentSheet.merge_cells(start_row=startRow+1,end_row=endRow,start_column=colNum,end_column=colNum)
                currentSheet.cell(row=startRow+1,column=colNum).alignment= Alignment(horizontal="center", vertical="center")

    return currentSheet.max_row


def processExecute(vendors,inputStartDate,inputEndDate,**keywords):
    startUnPRow = 1
    startMissRow =1

    WBs = ExcelUtilities.loadWorbook()
    unprocessedWB = WBs.unprocessed
    missingWB = WBs.missing

    for datasource in vendors['data']:
        keywords['cadence'] = datasource.get('cadence')
        vendor = keywords.get('vendor')
        for args in datasource.get('metadata'):

            if args.get('arrival_start_date') :
                arrival_date = args.get('arrival_start_date')
                arrival_date = datetime.date(datetime.strptime(arrival_date,"%Y-%m-%d"))
                start_date = GeneralUtils.getStartDate(max(arrival_date,inputStartDate))
            else:
                start_date = GeneralUtils.getStartDate(max(inputStartDate,date(2017,1,1)))

            end_date = GeneralUtils.getEndDate(inputEndDate)

            missingDatesSet = []

            yearMonthsRangeList = GeneralUtils.getMonthsRange(start_date,end_date)

            keywords['type'] = args.get('type')
            keywords['country'] = args.get('country')
            keywords['regex'] = args.get('regex')

            for country in keywords.get('country'):
                RawAvailableDatesSet = GeneralUtils.namedtuple_with_defaults('RawAvailableDatesSet','general')
                CleanAvailableDatesSet = GeneralUtils.namedtuple_with_defaults('CleanAvailableDatesSet','general')

                UnprocessedDatesSet = GeneralUtils.namedtuple_with_defaults('UnprocessedDatesSet','general')

                RawAvailableDatesSet.general = set()

                CleanAvailableDatesSet.general = set()


                args['vendor'] = vendor
                response = dict()
                cumulativeResponse = []
                for rawInfo in args.get('raw'):
                    for yearMonth_prefix in yearMonthsRangeList:
                        rawBucket = rawInfo.split('/')[0]
                        raw = rawInfo.replace('/','%',1).split('%')[1]
                        subs_value = {'country' : country, 'year' : yearMonth_prefix.split('-')[0], 'month' : yearMonth_prefix.split('-')[1]}
                        rawPrefix = GeneralUtils.prefixBuilder(raw,**subs_value)
                        response = client.list_objects_v2(Bucket= rawBucket ,Prefix = rawPrefix,Delimiter = '/')
                        cumulativeResponse.append(S3Utilities.getFinalContentFromResponse(client, response , rawBucket))
                        S3Utilities.finalContentList = []
                    # finalContentList = []
                for response in cumulativeResponse:
                    AvailableDatesAndSourceDict = hpeRawFileHandler(response,keywords.get('regex'))

                    for source,dates in AvailableDatesAndSourceDict.items():
                        keywords['datasource'] = source
                        RawAvailableDatesSet.general.update(dates)

                        cumulativeResponse = []
                        for cleanInfo in args.get('clean').get(source):
                            for yearMonth_prefix in yearMonthsRangeList:
                                cleanBucket = cleanInfo.split('/')[0]
                                clean = cleanInfo.replace('/','%',1).split('%')[1]
                                subs_value = {'country' : country.lower(), 'year' : yearMonth_prefix.split('-')[0], 'month' : yearMonth_prefix.split('-')[1]}
                                cleanPrefix = GeneralUtils.prefixBuilder(clean,**subs_value)
                                response = client.list_objects_v2(Bucket= cleanBucket ,Prefix = cleanPrefix,Delimiter = '/')
                                # finalContentList = []
                                AvailableDatesSet = hpeCleanFileHandler(response)
                                CleanAvailableDatesSet.general.update(AvailableDatesSet.general)


                        UnprocessedDatesSet.general = RawAvailableDatesSet.general - CleanAvailableDatesSet.general

                        UnprocessedDatesSet.general = GeneralUtils.getFilteredDates(UnprocessedDatesSet.general,start_date,end_date)

                        currentSheet = unprocessedWB[vendor]
                        #check if unprocessed dates set has only 1 element
                        keywords['type'] = args.get('type')
                        startUnPRow = hpeRowWriter(currentSheet,sorted(UnprocessedDatesSet.general),startUnPRow,**keywords)

                        unprocessedWB.save(unprocessedWB_out)

                        if datasource.get('cadence') == "daily":

                            missingDatesSet = set(GeneralUtils.d_range(start_date,end_date)) - (RawAvailableDatesSet.general.union(CleanAvailableDatesSet.general))

                        if datasource.get('cadence') == 'weekly':
                            missingDatesSet = []
                            weeks_set = set(GeneralUtils.w_range(start_date=start_date,end_date=end_date))

                            rawAvailableWeeksSet = GeneralUtils.getWeeksSet(RawAvailableDatesSet.general)
                            cleanAvailableWeeksSet = GeneralUtils.getWeeksSet(CleanAvailableDatesSet.general)
                            missingWeeksSet = weeks_set - (rawAvailableWeeksSet.union(cleanAvailableWeeksSet))

                            for missingWeeks in missingWeeksSet:
                                missingDatesSet.append(Week.day(missingWeeks,0))


                        if datasource.get('cadence') == 'monthly':
                            missingDatesSet =[]
                            monthsRange = GeneralUtils.getMonthsRange(start_date=start_date,end_date=end_date)

                            availableMonthsList = GeneralUtils.getMonthsSet(RawAvailableDatesSet.general.union(CleanAvailableDatesSet.general))
                            missingMonthsSet = set(monthsRange) - set(availableMonthsList)
                            for yearMonth in missingMonthsSet:
                                missingDate = yearMonth + '-01'
                                missingDate = datetime.date(datetime.strptime(missingDate,'%Y-%m-%d'))
                                missingDatesSet.append(missingDate)

                        currentSheet = missingWB[vendor]
                        keywords['type'] = args.get('type')
                        startMissRow = hpeRowWriter(currentSheet,sorted(missingDatesSet),startMissRow,**keywords)

                        missingWB.save(missingWB_out)


            print("Done ......." +datasource.get('datasource'))
