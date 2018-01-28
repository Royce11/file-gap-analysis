from com.reporting import S3Utilities,GeneralUtils,ExcelUtilities
from datetime import datetime,date
from isoweek import Week
from collections import OrderedDict
import openpyxl
from openpyxl.styles import NamedStyle,Font,Alignment

unprocessedWB_out = '/home/osboxes/shared-windows10/UnprocessedDates.xlsx'
missingWB_out = '/home/osboxes/shared-windows10/MissingDates.xlsx'
client = S3Utilities.getS3Client()

def rentrakFileHandler(response,dateRegex):
    generalFilePattern = []
    hispFilePattern = []

    FilePattern = GeneralUtils.namedtuple_with_defaults('FilePattern', 'general hispanic')

    for fullFileName in response:
        if "hisp" in fullFileName:
            hispanicExtractedDate=GeneralUtils.extractDate(dateRegex,fullFileName)
            if hispanicExtractedDate:
                hispFilePattern.append(GeneralUtils.getFormattedDate(hispanicExtractedDate))
        else:
            extractedDate=GeneralUtils.extractDate(dateRegex,fullFileName)
            if extractedDate:
                generalFilePattern.append(GeneralUtils.getFormattedDate(extractedDate))

    return FilePattern(set(generalFilePattern),set(hispFilePattern))

def rentrakRowWriter(currentSheet,datesSet,startRow,**keywords):
    tempDict = OrderedDict()

    currentSheet._current_row = startRow

    if len(datesSet):
        for dateElement in datesSet:
            tempDict.setdefault(str(dateElement.year) + '-' + str(dateElement.month),[]).append(dateElement.day)
        for key,value in tempDict.items():
            keywords['year-month'] = key
            keywords['dates'] = str(value)
            keywords['count'] = len(value)
            #vendor	datasource	cadence	type	country	year-month	dates	count
            fieldNamesDict = {'A': keywords.get('datasource'),
                              'B' : keywords.get('cadence'),
                              'C' : keywords.get('type'),
                              'D' : keywords.get('country'),
                              'E' : keywords.get('year-month'),
                              'F' : keywords.get('dates'),
                              'G' : keywords.get('count')}
            currentSheet.append(fieldNamesDict)

        endRow = currentSheet.max_row
        if currentSheet._current_row > 1:
            for colNum in range(1,5):
                currentSheet.merge_cells(start_row=startRow+1,end_row=endRow,start_column=colNum,end_column=colNum)
                currentSheet.cell(row=startRow+1,column=colNum).alignment= Alignment(horizontal="center", vertical="center")

    return currentSheet.max_row

def processExecute(vendors,inputStartDate,inputEndDate,**keywords):
    startUnPRow = 1
    startMissRow =1

    WBs = ExcelUtilities.loadWorbook()
    unprocessedWB = WBs.unprocessed
    missingWB = WBs.missing

    for datasource in vendors['data']:
        keywords['datasource'] = datasource.get('datasource')
        keywords['cadence'] = datasource.get('cadence')
        vendor = keywords.get('vendor')
        for args in datasource.get('metadata'):

            if args.get('arrival_start_date') :
                arrival_start_date = args.get('arrival_start_date')
                arrival_start_date = datetime.date(datetime.strptime(arrival_start_date,"%Y-%m-%d"))
                start_date = GeneralUtils.getStartDate(max(arrival_start_date,inputStartDate))
            else:
                start_date = GeneralUtils.getStartDate(inputStartDate)

            if args.get('arrival_end_date') :
                arrival_end_date = args.get('arrival_end_date')
                arrival_end_date = datetime.date(datetime.strptime(arrival_end_date,"%Y-%m-%d"))
                end_date = GeneralUtils.getEndDate(min(arrival_end_date,inputEndDate))
            else:
                end_date = GeneralUtils.getEndDate(inputEndDate)

            # end_date = GeneralUtils.getEndDate(inputEndDate)

            missingHispDatesSet =[]
            missingDatesSet = []

            yearMonthsRangeList = GeneralUtils.getMonthsRange(start_date,end_date,args.get('exclude_year_month'))

            start_date = datetime.date(datetime.strptime(min(yearMonthsRangeList) + '-01', '%Y-%m-%d'))
            end_date = datetime.date(datetime.strptime(max(yearMonthsRangeList) + '-30', '%Y-%m-%d'))

            country_list = args.get('country')
            keywords['regex'] = args.get('regex')

            for country in country_list:
                keywords['country'] = country
                RawAvailableDatesSet = GeneralUtils.namedtuple_with_defaults('RawAvailableDatesSet','general hispanic')
                CleanAvailableDatesSet = GeneralUtils.namedtuple_with_defaults('CleanAvailableDatesSet','general hispanic')

                AvailableDatesSet = GeneralUtils.namedtuple_with_defaults('AvailableDatesSet','general hispanic')

                UnprocessedDatesSet = GeneralUtils.namedtuple_with_defaults('UnprocessedDatesSet','general hispanic')
                MissedDatesSet = GeneralUtils.namedtuple_with_defaults('MissedDatesSet','general hispanic')

                RawAvailableDatesSet.general = set()
                RawAvailableDatesSet.hispanic = set()
                CleanAvailableDatesSet.general = set()
                CleanAvailableDatesSet.hispanic = set()


                args['vendor'] = vendor
                response = dict()
                cumulativeResponse = []
                for rawInfo in args.get('raw'):
                    for yearMonth_prefix in yearMonthsRangeList:
                        rawBucket = rawInfo.split('/')[0]
                        raw = rawInfo.replace('/','%',1).split('%')[1]
                        subs_value = {'country' : country, 'year' : yearMonth_prefix.split('-')[0], 'month' : yearMonth_prefix.split('-')[1]}
                        rawPrefix = GeneralUtils.prefixBuilder(raw,**subs_value)
                        response = client.list_objects_v2(Bucket= rawBucket ,Prefix = rawPrefix,Delimiter = '/')
                        cumulativeResponse.append(S3Utilities.getFinalContentFromResponse(client, response , rawBucket))
                        S3Utilities.finalContentList = []
                        # finalContentList = []
                flat_cumulativeResponse = [item for sublist in cumulativeResponse for item in sublist]
                AvailableDatesSet = rentrakFileHandler(flat_cumulativeResponse,keywords.get('regex'))
                if AvailableDatesSet.general:
                    RawAvailableDatesSet.general.update(AvailableDatesSet.general)
                if AvailableDatesSet.hispanic:
                    RawAvailableDatesSet.hispanic.update(AvailableDatesSet.hispanic)


                cumulativeResponse = []
                for cleanInfo in args.get('clean'):
                    for yearMonth_prefix in yearMonthsRangeList:
                        cleanBucket = cleanInfo.split('/')[0]
                        clean = cleanInfo.replace('/','%',1).split('%')[1]
                        subs_value = {'country' : country.lower(), 'year' : yearMonth_prefix.split('-')[0], 'month' : yearMonth_prefix.split('-')[1]}
                        cleanPrefix = GeneralUtils.prefixBuilder(clean,**subs_value)
                        response = client.list_objects_v2(Bucket= cleanBucket ,Prefix = cleanPrefix,Delimiter = '/')
                        cumulativeResponse.append(S3Utilities.getFinalContentFromResponse(client, response , cleanBucket))
                        S3Utilities.finalContentList = []
                        # finalContentList = []
                flat_cumulativeResponse = [item for sublist in cumulativeResponse for item in sublist]
                AvailableDatesSet = rentrakFileHandler(flat_cumulativeResponse,keywords.get('regex'))

                if AvailableDatesSet.general:
                    CleanAvailableDatesSet.general.update(AvailableDatesSet.general)
                if AvailableDatesSet.hispanic:
                    CleanAvailableDatesSet.hispanic.update(AvailableDatesSet.hispanic)


                UnprocessedDatesSet.general = RawAvailableDatesSet.general - CleanAvailableDatesSet.general
                UnprocessedDatesSet.general = GeneralUtils.getFilteredDates(UnprocessedDatesSet.general,start_date,end_date)
                if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                    UnprocessedDatesSet.hispanic = RawAvailableDatesSet.hispanic - CleanAvailableDatesSet.hispanic
                    UnprocessedDatesSet.hispanic = GeneralUtils.getFilteredDates(UnprocessedDatesSet.hispanic,start_date,end_date)


                currentSheet = unprocessedWB[vendor]
                #check if unprocessed dates set has only 1 element
                if RawAvailableDatesSet.general or CleanAvailableDatesSet.general:
                    if UnprocessedDatesSet.general:
                        keywords['type'] = 'non-hispanic'
                        startUnPRow = rentrakRowWriter(currentSheet,sorted(UnprocessedDatesSet.general),startUnPRow,**keywords)
                # unprocessedWB.save(unprocessedWB_out)
                if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                    if UnprocessedDatesSet.hispanic:
                        keywords['type'] = 'hispanic'
                        startUnPRow = rentrakRowWriter(currentSheet,sorted(UnprocessedDatesSet.hispanic),startUnPRow,**keywords)

                unprocessedWB.save(unprocessedWB_out)

                if datasource.get('cadence') == "daily":
                    if RawAvailableDatesSet.general or CleanAvailableDatesSet.general:
                        missingDatesSet = set(GeneralUtils.d_range(start_date,end_date,datasource.get('cadence'))) - (RawAvailableDatesSet.general.union(CleanAvailableDatesSet.general))
                    if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                        missingHispDatesSet = set(GeneralUtils.d_range(date(start_date,end_date,datasource.get('cadence')))) - (RawAvailableDatesSet.hispanic.union(CleanAvailableDatesSet.hispanic))

                if datasource.get('cadence') == 'weekly':
                    missingDatesSet = []
                    if RawAvailableDatesSet.general or CleanAvailableDatesSet.general:
                        weeks_set = set(GeneralUtils.w_range(start_date,end_date=end_date))
                        rawAvailableWeeksSet = GeneralUtils.getWeeksSet(RawAvailableDatesSet.general)
                        cleanAvailableWeeksSet = GeneralUtils.getWeeksSet(CleanAvailableDatesSet.general)
                        missingWeeksSet = weeks_set - (rawAvailableWeeksSet.union(cleanAvailableWeeksSet))

                        for missingWeeks in missingWeeksSet:
                            missingDatesSet.append(Week.day(missingWeeks,0))

                    if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                        weeks_set = set(GeneralUtils.w_range(start_date,end_date=end_date))
                        rawAvailableWeeksSet = GeneralUtils.getWeeksSet(RawAvailableDatesSet.hispanic)
                        cleanAvailableWeeksSet = GeneralUtils.getWeeksSet(CleanAvailableDatesSet.hispanic)
                        missingWeeksSet = weeks_set - (rawAvailableWeeksSet.union(cleanAvailableWeeksSet))

                        for missingWeeks in missingWeeksSet:
                            missingHispDatesSet.append(Week.day(missingWeeks,0))

                if datasource.get('cadence') == 'monthly':
                    missingDatesSet =[]
                    if RawAvailableDatesSet.general or CleanAvailableDatesSet.general:
                        monthsRange = GeneralUtils.getMonthsRange(start_date=start_date,end_date=end_date)
                        availableMonthsList = GeneralUtils.getMonthsSet(RawAvailableDatesSet.general.union(CleanAvailableDatesSet.general))
                        if args.get('exclude_year_month'):
                            missingMonthsSet = set(monthsRange) - set(availableMonthsList).union(args.get('exclude_year_month'))
                        else :
                            missingMonthsSet = set(monthsRange) - set(availableMonthsList)
                        for yearMonth in missingMonthsSet:
                            missingDate = yearMonth + '-01'
                            missingDate = datetime.date(datetime.strptime(missingDate,'%Y-%m-%d'))
                            missingDatesSet.append(missingDate)
                    if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                        monthsRange = GeneralUtils.getMonthsRange(start_date=start_date,end_date=end_date)
                        availableMonthsList = GeneralUtils.getMonthsSet(RawAvailableDatesSet.hispanic.union(CleanAvailableDatesSet.hispanic))
                        if args.get('exclude_year_month'):
                            missingMonthsSet = set(monthsRange) - set(availableMonthsList).union(args.get('exclude_year_month'))
                        else :
                            missingMonthsSet = set(monthsRange) - set(availableMonthsList)
                        for yearMonth in missingMonthsSet:
                            missingDate = yearMonth + '-01'
                            missingDate = datetime.date(datetime.strptime(missingDate,'%Y-%m-%d'))
                            missingHispDatesSet.append(missingDate)

                currentSheet = missingWB[vendor]
                if missingDatesSet:
                    keywords['type'] = 'non-hispanic'
                    startMissRow = rentrakRowWriter(currentSheet,sorted(missingDatesSet),startMissRow,**keywords)

                    missingWB.save(missingWB_out)

                if missingHispDatesSet:
                    keywords['type'] = 'hispanic'
                    startMissRow = rentrakRowWriter(currentSheet,sorted(missingHispDatesSet),startMissRow,**keywords)

                    missingWB.save(missingWB_out)

        print("Done ......." +datasource.get('datasource'))
