import boto3
from com.reportFromJson import ComscoreIntlTransform
import json
from collections import OrderedDict
from datetime import datetime,timedelta,date
from isoweek import Week
from dateutil.rrule import rrule, MONTHLY
import openpyxl

client = boto3.client('s3')

unprocessedWB_out = '/home/osboxes/shared-windows10/UnprocessedDates.xlsx'
missingWB_out = '/home/osboxes/shared-windows10/MissingDates.xlsx'

def getPrefix(commonPrefixList):
    prefixList = []
    for prefixDict in commonPrefixList:
        prefixList.append(prefixDict.get('Prefix'))

    return prefixList

def getFinalContentFromResponse(response,bucket):
    if "CommonPrefixes" in response.keys():
        prefixList = getPrefix(response.get('CommonPrefixes'))
        for prefix in prefixList:
            response = client.list_objects_v2(Bucket=bucket,Prefix = prefix,Delimiter = '/')
            getFinalContentFromResponse(response,bucket)
    if "Contents" in response.keys():
        for content in response.get('Contents'):
            finalContentList.append(content.get('Key'))

    return finalContentList

def d_range(d1,d2,key):
    delta_param = {"daily" : 1, "weekly" : 7, "monthly" : 29}
    delta = d2 - d1 #assumes second date is always after first
    modif_delta = int(delta.days)/delta_param.get(key)
    return [d1 + timedelta(days=i*delta_param.get(key)) for i in range(int(modif_delta) + 1)]

def getWeekFromDate(dateElement):
    return Week.withdate(dateElement)

def getWeeksSet(datesList):
    weeksList = []
    if datesList:
        for dateElement in datesList:
            weeksList.append(getWeekFromDate(dateElement))
    return set(weeksList)

def w_range(start_date):
    weeks_list = []
    startWeek = Week.withdate(start_date)
    endWeek = Week.thisweek()
    int_week = startWeek
    while int_week < endWeek:
        int_week = int_week + 1
        weeks_list.append(int_week)
    return weeks_list

def getMonthsSet(datesList):
    monthsList = []
    if datesList:
        for dateElement in datesList:
            if dateElement.day > 16:
                dateElement = dateElement+timedelta(days=15)
            monthsList.append(datetime.strftime(dateElement,'%Y-%m'))
    return monthsList

def getMonthsRange(start_date,end_date=None):
    monthsList = []
    if end_date is None:
        end_date = date.today()
    for dt in rrule(MONTHLY, dtstart=start_date, until=end_date):
        monthsList.append(datetime.strftime(dt.date(),'%Y-%m'))
    return monthsList

jsonObj = open('/home/osboxes/shared-windows10/comscore_try.json','r')
comscore_dict = json.load(jsonObj)

def initExcelWB():
    wb1 = openpyxl.Workbook()
    wb1.save(unprocessedWB_out)
    wb2 = openpyxl.Workbook()
    wb2.save(missingWB_out)

initExcelWB()

def initExcelSheet(sheetTitle,fieldnames):
    missingWB = openpyxl.load_workbook(missingWB_out)
    unprocessedWB = openpyxl.load_workbook(unprocessedWB_out)
    missingWB.create_sheet(sheetTitle)
    unprocessedWB.create_sheet(sheetTitle)
    currentSheet1 = missingWB.get_sheet_by_name(sheetTitle)
    currentSheet2 = unprocessedWB.get_sheet_by_name(sheetTitle)
    i=0
    while i < len(fieldnames):
        currentSheet1.cell(row=1,column=i+1,value=fieldnames[i])
        currentSheet2.cell(row=1,column=i+1,value=fieldnames[i])
        i=i+1
    missingWB.save(missingWB_out)
    unprocessedWB.save(unprocessedWB_out)

def getPrefixArray(rawPrefixes):
    rawPrefixList = []
    for rawPrefix in rawPrefixes:
        rawPrefixList.append(rawPrefix)


def rowWriter(currentSheet,datesSet,**keywords):
    tempDict = OrderedDict()

    if len(datesSet):
        for dateElement in datesSet:
            tempDict.setdefault(str(dateElement.year) + '-' + str(dateElement.month),[]).append(dateElement.day)
        for key,value in tempDict.items():
            keywords['year-month'] = key
            keywords['dates'] = str(value)
            keywords['count'] = len(value)
            #vendor	datasource	cadence	type	country	year-month	dates	count
            fieldNamesDict = {'A': keywords.get('datasource'),
                              'B' : keywords.get('cadence'),
                              'C' : keywords.get('type'),
                              'D' : keywords.get('country'),
                              'E' : keywords.get('year-month'),
                              'F' : keywords.get('dates'),
                              'G' : keywords.get('count')}
            currentSheet.append(fieldNamesDict)


    return keywords


for vendors in comscore_dict:
    initExcelSheet(vendors.get('vendor'),vendors.get('fieldnames'))
    missingWB = openpyxl.load_workbook(missingWB_out)
    unprocessedWB = openpyxl.load_workbook(unprocessedWB_out)

    keywords = {'vendor': vendors.get('vendor'),'fieldnames' : vendors.get('fieldnames')}

    finalContentList = []
    compiled_List = []
    rowNum = 2
    for datasource in vendors['data']:
        keywords['datasource'] = datasource.get('datasource')
        keywords['cadence'] = datasource.get('cadence')
        keywords['rowNum'] = rowNum
        for args in datasource.get('metadata'):

            keywords['type'] = args.get('type')
            keywords['country'] = args.get('country')

            rawAvailableDatesSet = set()
            cleanAvailableDatesSet = set()

            args['cleanFlag'] = False
            response = dict()
            cumulativeResponse = []
            for rawInfo in args.get('raw'):
                rawBucket = rawInfo.split('/')[0]
                rawPrefix = rawInfo.replace('/','%',1).split('%')[1]
                response = client.list_objects_v2(Bucket= rawBucket ,Prefix = rawPrefix,Delimiter = '/')
                cumulativeResponse.append(getFinalContentFromResponse(response , rawBucket))
            for response in cumulativeResponse:
                rawAvailableDatesSet.update(ComscoreIntlTransform.getDictListForHPEOriginal(response,**args))

            #Keep it false for rentrak and hpe
            args['cleanFlag'] = True
            cumulativeResponse = []
            for cleanInfo in args.get('clean'):
                cleanBucket = cleanInfo.split('/')[0]
                cleanPrefix = cleanInfo.replace('/','%',1).split('%')[1]
                # response = client.list_objects_v2(Bucket= cleanBucket ,Prefix = cleanPrefix,Delimiter = '/')
                cumulativeResponse.append(client.list_objects_v2(Bucket= cleanBucket ,Prefix = cleanPrefix,Delimiter = '/'))
            for response in cumulativeResponse:
                cleanAvailableDatesSet.update(ComscoreIntlTransform.getDictListForHPEOriginal(response,**args))

            # unprocessedDatesSet = rawAvailableDatesSet.difference(cleanAvailableDatesSet)

            unprocessedDatesSet = rawAvailableDatesSet - cleanAvailableDatesSet
            compiled_List.append(unprocessedDatesSet)

            currentSheet = unprocessedWB.get_sheet_by_name(keywords.get('vendor'))
            #check if unprocessed dates set has only 1 element
            fieldNamesDict = rowWriter(currentSheet,sorted(unprocessedDatesSet),**keywords)

            unprocessedWB.save(unprocessedWB_out)

            if datasource.get('cadence') == "daily":
                missingDatesSet = set(d_range(date(2017,1,1),date.today(),datasource.get('cadence'))) - (rawAvailableDatesSet.union(cleanAvailableDatesSet))
                # compiled_List.append(missingDatesSet)

            if datasource.get('cadence') == 'weekly':
                missingDatesSet = []
                weeksRange =[]
                rawAvailableWeeksSet = getWeeksSet(rawAvailableDatesSet)
                cleanAvailableWeeksSet = getWeeksSet(cleanAvailableDatesSet)
                weeks_set = set(w_range(date(2017,1,1)))
                missingWeeksSet = weeks_set - (rawAvailableWeeksSet.union(cleanAvailableWeeksSet))

                for missingWeeks in missingWeeksSet:
                    missingDatesSet.append(Week.day(missingWeeks,0))

            if datasource.get('cadence') == 'monthly':
                missingDatesSet =[]
                availableMonthsList = getMonthsSet(rawAvailableDatesSet.union(cleanAvailableDatesSet))
                monthsRange = getMonthsRange(start_date=date(2017,1,1))
                missingMonthsSet = set(monthsRange) - set(availableMonthsList)
                for yearMonth in missingMonthsSet:
                    missingDate = yearMonth + '-01'
                    missingDate = datetime.date(datetime.strptime(missingDate,'%Y-%m-%d'))
                    missingDatesSet.append(missingDate)

            currentSheet = missingWB.get_sheet_by_name(keywords.get('vendor'))

            fieldNamesDict = rowWriter(currentSheet,sorted(missingDatesSet),**keywords)

            missingWB.save(missingWB_out)

        print("Done ......." +datasource.get('datasource'))
