import boto3
from com.report import ComscoreIntlTransform
import json
from collections import OrderedDict
from datetime import datetime,timedelta,date
from isoweek import Week
from dateutil.rrule import rrule, MONTHLY
import openpyxl
from openpyxl.styles import NamedStyle,Font,Alignment
import sys
from com.reporting import RentrakHandler

client = boto3.client('s3')

unprocessedWB_out = '/home/osboxes/shared-windows10/UnprocessedDates.xlsx'
missingWB_out = '/home/osboxes/shared-windows10/MissingDates.xlsx'

def getPrefix(commonPrefixList):
    prefixList = []
    for prefixDict in commonPrefixList:
        prefixList.append(prefixDict.get('Prefix'))

    return prefixList

def getFinalContentFromResponse(response,bucket):
    if "CommonPrefixes" in response.keys():
        prefixList = getPrefix(response.get('CommonPrefixes'))
        for prefix in prefixList:
            response = client.list_objects_v2(Bucket=bucket,Prefix = prefix,Delimiter = '/')
            getFinalContentFromResponse(response,bucket)
    if "Contents" in response.keys():
        for content in response.get('Contents'):
            finalContentList.append(content.get('Key'))

    return finalContentList

def d_range(d1,d2,key):
    delta_param = {"daily" : 1, "weekly" : 7, "monthly" : 29}
    delta = d2 - d1 #assumes second date is always after first
    modif_delta = int(delta.days)/delta_param.get(key)
    return [d1 + timedelta(days=i*delta_param.get(key)) for i in range(int(modif_delta) + 1)]

def getWeekFromDate(dateElement):
    return Week.withdate(dateElement)

def getWeeksSet(datesList):
    weeksList = []
    if datesList:
        for dateElement in datesList:
            weeksList.append(getWeekFromDate(dateElement))
    return set(weeksList)

def w_range(start_date,end_date=None):
    weeks_list = []
    startWeek = Week.withdate(start_date)
    if end_date:
        endWeek = Week.withdate(end_date)
    else:
        endWeek = Week.thisweek()
    int_week = startWeek
    while int_week < endWeek:
        int_week = int_week + 1
        weeks_list.append(int_week)
    return weeks_list

def getMonthsSet(datesList):
    monthsList = []
    if datesList:
        for dateElement in datesList:
            if dateElement.day > 16:
                dateElement = dateElement+timedelta(days=15)
            monthsList.append(datetime.strftime(dateElement,'%Y-%m'))
    return monthsList

def getMonthsRange(start_date,end_date=None):
    monthsList = []
    if end_date is None:
        end_date = date.today()
    for dt in rrule(MONTHLY, dtstart=start_date, until=end_date):
        monthsList.append(datetime.strftime(dt.date(),'%Y-%m'))
    return monthsList

jsonObj = open('/home/osboxes/shared-windows10/rentrak_try.json','r')
comscore_dict = json.load(jsonObj)

def initExcelWB():
    wb1 = openpyxl.Workbook()
    wb1.save(unprocessedWB_out)
    wb2 = openpyxl.Workbook()
    wb2.save(missingWB_out)

initExcelWB()

def addHeaderStyle(sheetTitle):
    missingWB = openpyxl.load_workbook(missingWB_out)
    unprocessedWB = openpyxl.load_workbook(unprocessedWB_out)
    currentSheet1 = missingWB.get_sheet_by_name(sheetTitle)
    currentSheet2 = unprocessedWB.get_sheet_by_name(sheetTitle)
    if 'red_bold' not in missingWB._named_styles:
        red_bold = NamedStyle(name="red_bold")
    red_bold.font = Font(color='00FF0000', bold=True)
    missingWB.add_named_style(red_bold)

    if 'red_bold' not in unprocessedWB._named_styles:
        red_bold = NamedStyle(name="red_bold")
    red_bold.font = Font(color='00FF0000', bold=True)
    unprocessedWB.add_named_style(red_bold)

    for cell in currentSheet1["1:1"]:
        cell.style = 'red_bold'
    for cell in currentSheet2["1:1"]:
        cell.style = 'red_bold'

    missingWB.save(missingWB_out)
    unprocessedWB.save(unprocessedWB_out)

def initExcelSheet(sheetTitle,fieldnames):
    missingWB = openpyxl.load_workbook(missingWB_out)
    unprocessedWB = openpyxl.load_workbook(unprocessedWB_out)
    missingWB.create_sheet(sheetTitle)
    unprocessedWB.create_sheet(sheetTitle)
    currentSheet1 = missingWB.get_sheet_by_name(sheetTitle)
    currentSheet2 = unprocessedWB.get_sheet_by_name(sheetTitle)

    i=0
    while i < len(fieldnames):
        currentSheet1.cell(row=1,column=i+1,value=fieldnames[i])

        currentSheet2.cell(row=1,column=i+1,value=fieldnames[i])
        i=i+1
    missingWB.save(missingWB_out)
    unprocessedWB.save(unprocessedWB_out)

    addHeaderStyle(sheetTitle)

def getPrefixArray(rawPrefixes):
    rawPrefixList = []
    for rawPrefix in rawPrefixes:
        rawPrefixList.append(rawPrefix)


def rowWriter(currentSheet,datesSet,startRow,**keywords):
    tempDict = OrderedDict()

    currentSheet._current_row = startRow

    if len(datesSet):
        for dateElement in datesSet:
            tempDict.setdefault(str(dateElement.year) + '-' + str(dateElement.month),[]).append(dateElement.day)
        for key,value in tempDict.items():
            keywords['year-month'] = key
            keywords['dates'] = str(value)
            keywords['count'] = len(value)
            #vendor	datasource	cadence	type	country	year-month	dates	count
            fieldNamesDict = {'A': keywords.get('datasource'),
                              'B' : keywords.get('cadence'),
                              'C' : keywords.get('type'),
                              'D' : keywords.get('country'),
                              'E' : keywords.get('year-month'),
                              'F' : keywords.get('dates'),
                              'G' : keywords.get('count')}
            currentSheet.append(fieldNamesDict)

        endRow = currentSheet.max_row
        if currentSheet.max_row > 1:
            for colNum in range(1,5):
                currentSheet.merge_cells(start_row=startRow+1,end_row=endRow,start_column=colNum,end_column=colNum)
                currentSheet.cell(row=startRow+1,column=colNum).alignment= Alignment(horizontal="center", vertical="center")

    return currentSheet.max_row

def getStartDate(dateElement):
    dateElement = datetime.date(datetime.strptime(dateElement,'%Y-%m-%d'))
    return max(dateElement,date(2016,1,1))

def getEndDate(dateElement):
    dateElement = datetime.date(datetime.strptime(dateElement,'%Y-%m-%d'))
    return min(dateElement,date.today())

def getHandler(vendors, **keywords):
    vendor = keywords.get('vendor')
    if vendor == 'rentrak':
        return RentrakHandler(vendors,**keywords)

# for vendors in comscore_dict:
#     initExcelSheet(vendors.get('vendor'),vendors.get('fieldnames'))
#     missingWB = openpyxl.load_workbook(missingWB_out)
#     unprocessedWB = openpyxl.load_workbook(unprocessedWB_out)
#
#     keywords = {'vendor': vendors.get('vendor'),'fieldnames' : vendors.get('fieldnames')}
#
#     finalContentList = []
#     compiled_List = []
#     startUnPRow = 1
#     startMissRow =1
#     datasourceCadenceStartRow = 1
#
#     getHandler(**keywords)
#
#     for datasource in vendors['data']:
#         keywords['datasource'] = datasource.get('datasource')
#         keywords['cadence'] = datasource.get('cadence')
#
#         # currentSheetUnP = unprocessedWB.get_sheet_by_name(keywords.get('vendor'))
#         # if currentSheetUnP.max_row > 1:
#         #     currentSheet.merge_cells(start_row=startUnPRow+1,end_row=currentSheetUnP.max_row,start_column=1,end_column=2)
#         #     unprocessedWB.save(unprocessedWB_out)
#         #
#         #
#         # currentSheetMiss = unprocessedWB.get_sheet_by_name(keywords.get('vendor'))
#         # if currentSheetMiss.max_row > 1:
#         #     currentSheet.merge_cells(start_row=startMissRow+1,end_row=currentSheetMiss.max_row,start_column=1,end_column=2)
#         #     missingWB.save(missingWB_out)
#
#
#
#         for args in datasource.get('metadata'):
#
#             if args.get('arrival_start_date') :
#                 start_date = getStartDate(args.get('arrival_start_date'))
#             else:
#                 start_date = date(2016,1,1)
#
#
#             missingHispDatesSet =[]
#             missingDatesSet = []
#
#             keywords['type'] = args.get('type')
#             keywords['country'] = args.get('country')
#
#             RawAvailableDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('RawAvailableDatesSet','general hispanic')
#             CleanAvailableDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('CleanAvailableDatesSet','general hispanic')
#
#             AvailableDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('AvailableDatesSet','general hispanic')
#
#             UnprocessedDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('UnprocessedDatesSet','general hispanic')
#             MissedDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('MissedDatesSet','general hispanic')
#
#             RawAvailableDatesSet.general = set()
#             RawAvailableDatesSet.hispanic = set()
#             CleanAvailableDatesSet.general = set()
#             CleanAvailableDatesSet.hispanic = set()
#
#
#             args['cleanFlag'] = False
#             args['vendor'] = keywords.get('vendor')
#             response = dict()
#             cumulativeResponse = []
#             for rawInfo in args.get('raw'):
#                 rawBucket = rawInfo.split('/')[0]
#                 rawPrefix = rawInfo.replace('/','%',1).split('%')[1]
#                 response = client.list_objects_v2(Bucket= rawBucket ,Prefix = rawPrefix,Delimiter = '/')
#                 cumulativeResponse.append(getFinalContentFromResponse(response , rawBucket))
#                 finalContentList = []
#             for response in cumulativeResponse:
#                 AvailableDatesSet = ComscoreIntlTransform.getDictListForHPEOriginal(response,**args)
#                 RawAvailableDatesSet.general.update(AvailableDatesSet.general)
#                 if keywords.get('vendor') == 'rentrak':
#                     RawAvailableDatesSet.hispanic.update(AvailableDatesSet.hispanic)
#
#
#             #Keep it false for rentrak and hpe
#             args['cleanFlag'] = False
#             cumulativeResponse = []
#             for cleanInfo in args.get('clean'):
#                 cleanBucket = cleanInfo.split('/')[0]
#                 cleanPrefix = cleanInfo.replace('/','%',1).split('%')[1]
#                 response = client.list_objects_v2(Bucket= cleanBucket ,Prefix = cleanPrefix,Delimiter = '/')
#                 cumulativeResponse.append(getFinalContentFromResponse(response , cleanBucket))
#                 finalContentList = []
#             for response in cumulativeResponse:
#                 AvailableDatesSet = ComscoreIntlTransform.getDictListForHPEOriginal(response,**args)
#                 CleanAvailableDatesSet.general.update(AvailableDatesSet.general)
#                 if CleanAvailableDatesSet.hispanic:
#                     CleanAvailableDatesSet.hispanic.update(AvailableDatesSet.hispanic)
#
#
#             UnprocessedDatesSet.general = RawAvailableDatesSet.general - CleanAvailableDatesSet.general
#             if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
#                 UnprocessedDatesSet.hispanic = RawAvailableDatesSet.hispanic - CleanAvailableDatesSet.hispanic
#
#             currentSheet = unprocessedWB.get_sheet_by_name(keywords.get('vendor'))
#             #check if unprocessed dates set has only 1 element
#             keywords['type'] = args.get('type')
#             startUnPRow = rowWriter(currentSheet,sorted(UnprocessedDatesSet.general),startUnPRow,**keywords)
#             # unprocessedWB.save(unprocessedWB_out)
#             if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
#                 if UnprocessedDatesSet.hispanic:
#                     keywords['type'] = 'hispanic'
#                     startUnPRow = rowWriter(currentSheet,sorted(UnprocessedDatesSet.hispanic),startUnPRow,**keywords)
#
#             unprocessedWB.save(unprocessedWB_out)
#
#             if datasource.get('cadence') == "daily":
#
#                 missingDatesSet = set(d_range(start_date,date.today(),datasource.get('cadence'))) - (RawAvailableDatesSet.general.union(CleanAvailableDatesSet.general))
#                 if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
#                     missingHispDatesSet = set(d_range(date(2016,1,1),date.today(),datasource.get('cadence'))) - (RawAvailableDatesSet.hispanic.union(CleanAvailableDatesSet.hispanic))
#
#             if datasource.get('cadence') == 'weekly':
#                 missingDatesSet = []
#                 weeksRange =[]
#                 rawAvailableWeeksSet = getWeeksSet(RawAvailableDatesSet.general)
#                 cleanAvailableWeeksSet = getWeeksSet(CleanAvailableDatesSet.general)
#                 weeks_set = set(w_range(start_date))
#                 missingWeeksSet = weeks_set - (rawAvailableWeeksSet.union(cleanAvailableWeeksSet))
#
#                 for missingWeeks in missingWeeksSet:
#                     missingDatesSet.append(Week.day(missingWeeks,0))
#
#                 if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
#                     rawAvailableWeeksSet = getWeeksSet(RawAvailableDatesSet.hispanic)
#                     cleanAvailableWeeksSet = getWeeksSet(CleanAvailableDatesSet.hispanic)
#                     missingWeeksSet = weeks_set - (rawAvailableWeeksSet.union(cleanAvailableWeeksSet))
#
#                     for missingWeeks in missingWeeksSet:
#                         missingHispDatesSet.append(Week.day(missingWeeks,0))
#
#             if datasource.get('cadence') == 'monthly':
#                 missingDatesSet =[]
#                 availableMonthsList = getMonthsSet(RawAvailableDatesSet.general.union(CleanAvailableDatesSet.general))
#                 monthsRange = getMonthsRange(start_date=start_date)
#                 missingMonthsSet = set(monthsRange) - set(availableMonthsList)
#                 for yearMonth in missingMonthsSet:
#                     missingDate = yearMonth + '-01'
#                     missingDate = datetime.date(datetime.strptime(missingDate,'%Y-%m-%d'))
#                     missingDatesSet.append(missingDate)
#                 if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
#                     availableMonthsList = getMonthsSet(RawAvailableDatesSet.hispanic.union(CleanAvailableDatesSet.hispanic))
#                     missingMonthsSet = set(monthsRange) - set(availableMonthsList)
#                     for yearMonth in missingMonthsSet:
#                         missingDate = yearMonth + '-01'
#                         missingDate = datetime.date(datetime.strptime(missingDate,'%Y-%m-%d'))
#                         missingHispDatesSet.append(missingDate)
#
#             currentSheet = missingWB.get_sheet_by_name(keywords.get('vendor'))
#             keywords['type'] = args.get('type')
#             startMissRow = rowWriter(currentSheet,sorted(missingDatesSet),startMissRow,**keywords)
#
#             missingWB.save(missingWB_out)
#
#             if missingHispDatesSet:
#                 keywords['type'] = 'hispanic'
#                 startMissRow = rowWriter(currentSheet,sorted(missingHispDatesSet),startMissRow,**keywords)
#
#                 missingWB.save(missingWB_out)
#
#         print("Done ......." +datasource.get('datasource'))


def rentrakHandler():
    startUnPRow = 1
    startMissRow =1
    for datasource in vendors['data']:
        keywords['datasource'] = datasource.get('datasource')
        keywords['cadence'] = datasource.get('cadence')

        for args in datasource.get('metadata'):

            if args.get('arrival_start_date') :
                start_date = getStartDate(args.get('arrival_start_date'))
            else:
                start_date = date(2016,1,1)


            missingHispDatesSet =[]
            missingDatesSet = []

            keywords['type'] = args.get('type')
            keywords['country'] = args.get('country')

            RawAvailableDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('RawAvailableDatesSet','general hispanic')
            CleanAvailableDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('CleanAvailableDatesSet','general hispanic')

            AvailableDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('AvailableDatesSet','general hispanic')

            UnprocessedDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('UnprocessedDatesSet','general hispanic')
            MissedDatesSet = ComscoreIntlTransform.namedtuple_with_defaults('MissedDatesSet','general hispanic')

            RawAvailableDatesSet.general = set()
            RawAvailableDatesSet.hispanic = set()
            CleanAvailableDatesSet.general = set()
            CleanAvailableDatesSet.hispanic = set()


            args['cleanFlag'] = False
            args['vendor'] = keywords.get('vendor')
            response = dict()
            cumulativeResponse = []
            for rawInfo in args.get('raw'):
                rawBucket = rawInfo.split('/')[0]
                rawPrefix = rawInfo.replace('/','%',1).split('%')[1]
                response = client.list_objects_v2(Bucket= rawBucket ,Prefix = rawPrefix,Delimiter = '/')
                cumulativeResponse.append(getFinalContentFromResponse(response , rawBucket))
                finalContentList = []
            for response in cumulativeResponse:
                AvailableDatesSet = ComscoreIntlTransform.getDictListForHPEOriginal(response,**args)
                RawAvailableDatesSet.general.update(AvailableDatesSet.general)
                if keywords.get('vendor') == 'rentrak':
                    RawAvailableDatesSet.hispanic.update(AvailableDatesSet.hispanic)


            #Keep it false for rentrak and hpe
            args['cleanFlag'] = False
            cumulativeResponse = []
            for cleanInfo in args.get('clean'):
                cleanBucket = cleanInfo.split('/')[0]
                cleanPrefix = cleanInfo.replace('/','%',1).split('%')[1]
                response = client.list_objects_v2(Bucket= cleanBucket ,Prefix = cleanPrefix,Delimiter = '/')
                cumulativeResponse.append(getFinalContentFromResponse(response , cleanBucket))
                finalContentList = []
            for response in cumulativeResponse:
                AvailableDatesSet = ComscoreIntlTransform.getDictListForHPEOriginal(response,**args)
                CleanAvailableDatesSet.general.update(AvailableDatesSet.general)
                if CleanAvailableDatesSet.hispanic:
                    CleanAvailableDatesSet.hispanic.update(AvailableDatesSet.hispanic)


            UnprocessedDatesSet.general = RawAvailableDatesSet.general - CleanAvailableDatesSet.general
            if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                UnprocessedDatesSet.hispanic = RawAvailableDatesSet.hispanic - CleanAvailableDatesSet.hispanic

            currentSheet = unprocessedWB.get_sheet_by_name(keywords.get('vendor'))
            #check if unprocessed dates set has only 1 element
            keywords['type'] = args.get('type')
            startUnPRow = rowWriter(currentSheet,sorted(UnprocessedDatesSet.general),startUnPRow,**keywords)
            # unprocessedWB.save(unprocessedWB_out)
            if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                if UnprocessedDatesSet.hispanic:
                    keywords['type'] = 'hispanic'
                    startUnPRow = rowWriter(currentSheet,sorted(UnprocessedDatesSet.hispanic),startUnPRow,**keywords)

            unprocessedWB.save(unprocessedWB_out)

            if datasource.get('cadence') == "daily":
                missingDatesSet = []
                missingDatesSet = set(d_range(start_date,date.today(),datasource.get('cadence'))) - (RawAvailableDatesSet.general.union(CleanAvailableDatesSet.general))
                if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                    missingHispDatesSet = set(d_range(date(2016,1,1),date.today(),datasource.get('cadence'))) - (RawAvailableDatesSet.hispanic.union(CleanAvailableDatesSet.hispanic))

            if datasource.get('cadence') == 'weekly':
                missingDatesSet = []
                weeksRange =[]
                rawAvailableWeeksSet = getWeeksSet(RawAvailableDatesSet.general)
                cleanAvailableWeeksSet = getWeeksSet(CleanAvailableDatesSet.general)
                weeks_set = set(w_range(start_date))
                missingWeeksSet = weeks_set - (rawAvailableWeeksSet.union(cleanAvailableWeeksSet))

                for missingWeeks in missingWeeksSet:
                    missingDatesSet.append(Week.day(missingWeeks,0))

                if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                    rawAvailableWeeksSet = getWeeksSet(RawAvailableDatesSet.hispanic)
                    cleanAvailableWeeksSet = getWeeksSet(CleanAvailableDatesSet.hispanic)
                    missingWeeksSet = weeks_set - (rawAvailableWeeksSet.union(cleanAvailableWeeksSet))

                    for missingWeeks in missingWeeksSet:
                        missingHispDatesSet.append(Week.day(missingWeeks,0))

            if datasource.get('cadence') == 'monthly':
                missingDatesSet =[]
                availableMonthsList = getMonthsSet(RawAvailableDatesSet.general.union(CleanAvailableDatesSet.general))
                monthsRange = getMonthsRange(start_date=start_date)
                missingMonthsSet = set(monthsRange) - set(availableMonthsList)
                for yearMonth in missingMonthsSet:
                    missingDate = yearMonth + '-01'
                    missingDate = datetime.date(datetime.strptime(missingDate,'%Y-%m-%d'))
                    missingDatesSet.append(missingDate)
                if RawAvailableDatesSet.hispanic or CleanAvailableDatesSet.hispanic:
                    availableMonthsList = getMonthsSet(RawAvailableDatesSet.hispanic.union(CleanAvailableDatesSet.hispanic))
                    missingMonthsSet = set(monthsRange) - set(availableMonthsList)
                    for yearMonth in missingMonthsSet:
                        missingDate = yearMonth + '-01'
                        missingDate = datetime.date(datetime.strptime(missingDate,'%Y-%m-%d'))
                        missingHispDatesSet.append(missingDate)

            currentSheet = missingWB.get_sheet_by_name(keywords.get('vendor'))
            keywords['type'] = args.get('type')
            startMissRow = rowWriter(currentSheet,sorted(missingDatesSet),startMissRow,**keywords)

            missingWB.save(missingWB_out)

            if missingHispDatesSet:
                keywords['type'] = 'hispanic'
                startMissRow = rowWriter(currentSheet,sorted(missingHispDatesSet),startMissRow,**keywords)

                missingWB.save(missingWB_out)

        print("Done ......." +datasource.get('datasource'))


for vendors in comscore_dict:
    initExcelSheet(vendors.get('vendor'),vendors.get('fieldnames'))
    missingWB = openpyxl.load_workbook(missingWB_out)
    unprocessedWB = openpyxl.load_workbook(unprocessedWB_out)

    keywords = {'vendor': vendors.get('vendor'),'fieldnames' : vendors.get('fieldnames')}

    finalContentList = []
    compiled_List = []

    datasourceCadenceStartRow = 1

    getHandler(vendors,**keywords)