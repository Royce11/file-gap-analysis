import openpyxl
import os

wb = openpyxl.Workbook()
wb.create_sheet('sheet1',0)
wb.create_sheet('sheet2',1)
sheet1 = wb.get_sheet_by_name('sheet1')
sheet2 = wb.get_sheet_by_name('sheet2')

d = {'test' : [0,1,2],'foo':[342]}
d['new', 'key'] = d.pop('test')

fieldnames = ["vendor","datasource","cadence","type","country","year-month","dates","count"]
i=1
while i <= len(fieldnames):
    sheet1.cell(row=1,column=i,value=fieldnames[i-1])
    sheet2.cell(row=2,column=i,value=fieldnames[i-1])
    i = i+1
print (os.path.abspath('.'))
wb.save('/home/osboxes/shared-windows10/maauun.xlsx')
sheet1.merge_cells('B2:D3')
sheet1._current_row = 3
# for colNum in range(1,5):
#     sheet1.merge_cells(start_row=1, end_row=3, start_column=colNum,end_column=colNum)
#     wb.save('/home/osboxes/shared-windows10/maauun.xlsx')
sheet1.append({'A':"ffsdf",'B' : "asfdsf"})
wb.save('/home/osboxes/shared-windows10/maauun.xlsx')
curr = sheet1._current_row
print(curr)

